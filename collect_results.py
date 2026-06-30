"""
실험 결과를 수집하여 Excel + PNG 그래프로 저장.

사용법:
  python collect_results.py \\
    --results-jsons ./results/pair_0.8B/eval_results.json \\
                    ./results/pair_2B/eval_results.json \\
    --output routing_results.xlsx

각 JSON 파일 형식 (evaluate.py --output-json 출력):
  {
    "weak_model": "Qwen/Qwen3.5-2B",
    "strong_model": "Qwen/Qwen3.5-9B",
    "weak_only_accuracy": 62.4,
    "strong_only_accuracy": 81.3,
    "results": [
      {"method": "random", "threshold": 0.1, "strong_percentage": 10.0, "accuracy": 65.2},
      ...
    ]
  }
"""

import argparse
import json
import os
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

# openpyxl is optional: the pipeline is expensive (GPU hours), so if it is not
# installed we must NOT crash at the very last step. We fall back to CSV output
# and still produce the PNG graph.
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────
# Load
# ─────────────────────────────────────────────

def load_result(json_path: str, embedding: str = "default") -> dict:
    with open(json_path) as f:
        data = json.load(f)
    weak = data["weak_model"].split("/")[-1]   # "Qwen3.5-2B"
    strong = data["strong_model"].split("/")[-1]
    label = f"{weak} vs {strong}"
    df = pd.DataFrame(data["results"])
    df["pair"] = label
    df["embedding"] = embedding
    df["weak_model"] = data["weak_model"]
    df["strong_model"] = data["strong_model"]
    return {
        "label": label,
        "embedding": embedding,
        "weak_acc": data["weak_only_accuracy"],
        "strong_acc": data["strong_only_accuracy"],
        "weak_model": data["weak_model"],
        "strong_model": data["strong_model"],
        "df": df,
    }


def _parse_entry(entry: str) -> tuple[str, str]:
    """'embedding=path' → (embedding, path). '=' 없으면 (default, path)."""
    if "=" in entry:
        emb, path = entry.split("=", 1)
        return emb, path
    return "default", entry


# ─────────────────────────────────────────────
# Graph
# ─────────────────────────────────────────────

METHOD_STYLE = {
    "random":   {"color": "#888888", "linestyle": "--", "linewidth": 1.5, "marker": None},
    "mf":       {"color": "#2196F3", "linestyle": "-",  "linewidth": 2.0, "marker": "o"},
    "uniroute": {"color": "#FF9800", "linestyle": "-",  "linewidth": 2.0, "marker": "s"},
    "permodel": {"color": "#4CAF50", "linestyle": "-",  "linewidth": 2.0, "marker": "^"},
}
METHOD_LABEL = {
    "random":   "Random",
    "mf":       "MF Router",
    "uniroute": "UniRoute (K-Means)",
    "permodel": "Per-model Regression",
}


# 임베딩별 선 스타일 (색=method, 선모양=embedding)
_EMB_LINESTYLES = ["-", "--", ":", "-."]


def make_graphs(entries: list[dict], output_png: str, include_random: bool = False) -> str:
    """
    pair마다 하나의 subplot. 한 subplot 안에 임베딩 × method 곡선을 겹쳐 그린다.
    색 = method, 선모양 = embedding (e5-small/e5-large…).
    weak/strong 기준선은 임베딩과 무관하므로 pair당 한 번만.
    random은 사실상 대각선 baseline이라 기본적으로 그래프에서 제외(데이터 표엔 유지).
    """
    methods = (["random"] if include_random else []) + ["mf", "uniroute", "permodel"]
    # pair 순서 보존
    pairs = list(dict.fromkeys(e["label"] for e in entries))
    embeddings = list(dict.fromkeys(e["embedding"] for e in entries))
    emb_ls = {emb: _EMB_LINESTYLES[i % len(_EMB_LINESTYLES)] for i, emb in enumerate(embeddings)}
    multi_emb = len(embeddings) > 1

    n = len(pairs)
    fig, axes = plt.subplots(1, n, figsize=(7 * n, 5.5), sharey=False)
    if n == 1:
        axes = [axes]

    for ax, pair in zip(axes, pairs):
        pair_entries = [e for e in entries if e["label"] == pair]
        weak_acc = pair_entries[0]["weak_acc"]
        strong_acc = pair_entries[0]["strong_acc"]

        for e in pair_entries:
            df = e["df"]
            ls = emb_ls[e["embedding"]]
            for method in methods:
                df_m = df[df["method"] == method].sort_values("strong_percentage")
                if df_m.empty:
                    continue
                style = METHOD_STYLE.get(method, {})
                lbl = METHOD_LABEL.get(method, method)
                if multi_emb:
                    lbl = f"{lbl} · {e['embedding']}"
                ax.plot(
                    df_m["strong_percentage"], df_m["accuracy"],
                    label=lbl,
                    color=style.get("color", "black"),
                    linestyle=ls,
                    linewidth=style.get("linewidth", 1.8),
                    marker=style.get("marker"),
                    markersize=4,
                )

        ax.axhline(weak_acc, color="#555555", linestyle=":", linewidth=1.0,
                   label=f"Weak only ({weak_acc:.1f}%)")
        ax.axhline(strong_acc, color="#C62828", linestyle=":", linewidth=1.0,
                   label=f"Strong only ({strong_acc:.1f}%)")

        ax.set_xlabel("Strong Model Calls (%)", fontsize=11)
        ax.set_ylabel("Pass Rate (%)", fontsize=11)
        ax.set_title(pair, fontsize=12, fontweight="bold")
        ax.legend(loc="lower right", fontsize=8)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-2, 102)

    plt.tight_layout()
    plt.savefig(output_png, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved graph → {output_png}")
    return output_png


# ─────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────

def _header_style(ws, row_idx: int, n_cols: int):
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def write_sheet1(wb, entries: list[dict]):
    ws = wb.create_sheet("Deferral Curves")
    headers = ["Pair", "Embedding", "Method", "Threshold", "Pass Rate (%)", "Weak (%)", "Strong (%)"]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    for e in entries:
        df = e["df"].sort_values(["method", "strong_percentage"])
        for _, row in df.iterrows():
            ws.append([
                e["label"],
                e["embedding"],
                row["method"],
                round(float(row["threshold"]), 4),
                round(float(row["accuracy"]), 2),
                round(100.0 - float(row["strong_percentage"]), 2),
                round(float(row["strong_percentage"]), 2),
            ])

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4


def write_sheet2(wb, entries: list[dict]):
    ws = wb.create_sheet("Summary")
    headers = ["Pair", "Embedding", "Weak Model", "Strong Model", "Weak-only (%)", "Strong-only (%)"]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    for e in entries:
        ws.append([
            e["label"],
            e["embedding"],
            e["weak_model"],
            e["strong_model"],
            round(e["weak_acc"], 2),
            round(e["strong_acc"], 2),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4


def write_sheet_graph(wb, png_path: str):
    ws = wb.create_sheet("Graphs")
    img = XLImage(png_path)
    img.anchor = "A1"
    ws.add_image(img)


# ─────────────────────────────────────────────
# CSV fallback (when openpyxl is unavailable)
# ─────────────────────────────────────────────

def write_csv_fallback(entries: list[dict], output_xlsx: str) -> list[str]:
    """openpyxl이 없을 때 Excel 대신 CSV 2개로 저장한다."""
    base = os.path.splitext(output_xlsx)[0]
    curves_path = f"{base}_deferral_curves.csv"
    summary_path = f"{base}_summary.csv"

    curve_rows = []
    for e in entries:
        df = e["df"].sort_values(["method", "strong_percentage"])
        for _, row in df.iterrows():
            curve_rows.append({
                "Pair": e["label"],
                "Embedding": e["embedding"],
                "Method": row["method"],
                "Threshold": round(float(row["threshold"]), 4),
                "Pass Rate (%)": round(float(row["accuracy"]), 2),
                "Weak (%)": round(100.0 - float(row["strong_percentage"]), 2),
                "Strong (%)": round(float(row["strong_percentage"]), 2),
            })
    pd.DataFrame(curve_rows).to_csv(curves_path, index=False)

    summary_rows = [{
        "Pair": e["label"],
        "Embedding": e["embedding"],
        "Weak Model": e["weak_model"],
        "Strong Model": e["strong_model"],
        "Weak-only (%)": round(e["weak_acc"], 2),
        "Strong-only (%)": round(e["strong_acc"], 2),
    } for e in entries]
    pd.DataFrame(summary_rows).to_csv(summary_path, index=False)

    return [curves_path, summary_path]


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Collect routing experiment results → Excel + graphs")
    parser.add_argument(
        "--results-jsons",
        nargs="+",
        required=True,
        help="evaluate.py --output-json 결과 파일들. 임베딩 비교 시 'e5-small=path' "
        "처럼 라벨을 붙이면 같은 그래프/시트에 임베딩별로 겹쳐 그린다. "
        "(라벨 없으면 'default')",
    )
    parser.add_argument("--output", default="routing_results.xlsx", help="Output Excel path")
    parser.add_argument("--graph-random", action="store_true",
                        help="그래프에 random baseline 곡선도 포함 (기본: 제외, 데이터 표엔 유지)")
    args = parser.parse_args()

    entries = [load_result(path, emb) for emb, path in (_parse_entry(x) for x in args.results_jsons)]

    # PNG graph
    output_dir = str(Path(args.output).parent)
    png_path = os.path.join(output_dir, "routing_curves.png")
    make_graphs(entries, png_path, include_random=args.graph_random)

    # Excel (or CSV fallback if openpyxl is unavailable)
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        write_sheet1(wb, entries)
        write_sheet2(wb, entries)
        write_sheet_graph(wb, png_path)

        wb.save(args.output)
        print(f"Saved Excel → {args.output}")
    else:
        csv_paths = write_csv_fallback(entries, args.output)
        print(
            "[warn] openpyxl is not installed — wrote CSV instead of Excel.\n"
            "       Install it with:  pip install openpyxl   (or  pip install -e \".[eval]\")"
        )
        for p in csv_paths:
            print(f"Saved CSV → {p}")
        print(f"Graph available at → {png_path}")


if __name__ == "__main__":
    main()
